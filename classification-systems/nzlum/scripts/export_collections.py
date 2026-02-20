#!/usr/bin/env python3
from rdflib import Graph, Namespace, URIRef
from rdflib.namespace import RDF
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

SKOS = Namespace("http://www.w3.org/2004/02/skos/core#")

def get_en(g, s, p):
    """Get English language object, fallback to any."""
    for o in g.objects(s, p):
        if getattr(o, "language", None) == "en":
            return str(o)
    o = next(g.objects(s, p), None)
    return str(o) if o else None

def notation(g, c):
    """Get notation of a concept."""
    n = next(g.objects(c, SKOS.notation), None)
    return str(n) if n else ""

def label(g, c):
    """Get prefLabel of a concept."""
    return get_en(g, c, SKOS.prefLabel) or ""

def definition(g, c):
    """Get definition of a concept."""
    return get_en(g, c, SKOS.definition) or ""

def usage_note(g, c):
    """Get usageNote of a concept."""
    return get_en(g, c, SKOS.usageNote) or ""

def scope_note(g, c):
    """Get scopeNote of a concept."""
    return get_en(g, c, SKOS.scopeNote) or ""

def get_collection_members(g, collection):
    """Get members of a collection via skos:member."""
    members = set()
    for member in g.objects(collection, SKOS.member):
        members.add(member)
    return members

def main():
    if len(sys.argv) < 5:
        print("Usage: export_collections.py INPUT.ttl SCHEME_IRI OUT.md OUT.xlsx", file=sys.stderr)
        sys.exit(2)

    in_ttl, scheme_iri, out_md, out_xlsx = sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4]

    g = Graph()
    g.parse(in_ttl, format="turtle")

    # Expand CURIE using graph's namespace bindings
    scheme = URIRef(scheme_iri)
    if ":" in scheme_iri and not scheme_iri.startswith("http"):
        prefix, local = scheme_iri.split(":", 1)
        for p, n in g.namespaces():
            if p == prefix:
                scheme = URIRef(str(n) + local)
                break

    # Find all collections in the scheme
    collections = [c for c in g.subjects(RDF.type, SKOS.Collection) 
                   if (c, SKOS.inScheme, scheme) in g]
    
    if not collections:
        raise SystemExit(f"No skos:Collection with skos:inScheme {scheme_iri} found in {in_ttl}")

    # Sort collections by label
    sorted_colls = sorted(collections, key=lambda c: label(g, c).lower())

    # Build rows: (collection_label, notation, term_label, defn, usage, scope)
    rows = []
    for coll in sorted_colls:
        coll_label = label(g, coll)
        members = get_collection_members(g, coll)
        # Sort members alphabetically by notation
        sorted_members = sorted(members, key=lambda m: notation(g, m).lower())
        
        for member in sorted_members:
            not_str = notation(g, member)
            term_label = label(g, member)
            defn = definition(g, member)
            usage = usage_note(g, member)
            scope = scope_note(g, member)
            rows.append((coll_label, not_str, term_label, defn, usage, scope))

    # --- Markdown output (only Collection, Notation, Term) ---
    md_lines = ["| Collection | Notation | Term |", "|---|---|---|"]
    prev_coll = None
    for coll_label, not_str, term_label, defn, usage, scope in rows:
        coll_out = "" if coll_label == prev_coll else coll_label
        md_lines.append(f"| {coll_out} | {not_str} | {term_label} |")
        prev_coll = coll_label

    with open(out_md, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines) + "\n")

    # --- XLSX output with merged cells ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Collections"
    ws.append(["Collection", "Notation", "PrefLabel", "Definition", "UsageNote", "ScopeNote"])

    # Format header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Write data rows
    start_row = 2
    for coll_label, not_str, term_label, defn, usage, scope in rows:
        ws.append([coll_label, not_str, term_label, defn, usage, scope])

    end_row = start_row + len(rows) - 1

    # Merge Collection column where values repeat
    if rows:
        r = start_row
        while r <= end_row:
            v = ws.cell(row=r, column=1).value
            r2 = r
            while r2 + 1 <= end_row and ws.cell(row=r2 + 1, column=1).value == v:
                r2 += 1
            if v and r2 > r:
                ws.merge_cells(start_row=r, start_column=1, end_row=r2, end_column=1)
                ws.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
            r = r2 + 1

    # Column widths and text wrapping
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 35

    for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(out_xlsx)
    print(f"Wrote {out_md} and {out_xlsx}", file=sys.stderr)

if __name__ == "__main__":
    main()
