#!/usr/bin/env python3
from rdflib import Graph, Namespace, URIRef
from rdflib.namespace import RDF
from collections import defaultdict
import re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SKOS = Namespace("http://www.w3.org/2004/02/skos/core#")

def get_en(g, s, p):
    for o in g.objects(s, p):
        if getattr(o, "language", None) == "en":
            return str(o)
    o = next(g.objects(s, p), None)
    return str(o) if o else None

def notation(g, c):
    n = next(g.objects(c, SKOS.notation), None)
    return str(n) if n else ""

def label(g, c):
    return get_en(g, c, SKOS.prefLabel) or str(c)

def code_key(code: str):
    # numeric-aware sort for dotted numeric notations
    try:
        return [int(x) for x in code.split(".")]
    except Exception:
        return [code]

def main():
    if len(sys.argv) < 5:
        print("Usage: export_table.py INPUT.ttl SCHEME_IRI OUT.md OUT.xlsx", file=sys.stderr)
        sys.exit(2)

    in_ttl, scheme_iri, out_md, out_xlsx = sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4]
    scheme = URIRef(scheme_iri)

    g = Graph()
    g.parse(in_ttl, format="turtle")

    concepts = {c for c in g.subjects(RDF.type, SKOS.Concept) if (c, SKOS.inScheme, scheme) in g}

    tops = [c for c in g.objects(scheme, SKOS.hasTopConcept) if c in concepts]
    if not tops:
        raise SystemExit(f"No skos:hasTopConcept found for scheme {scheme_iri} (or top concepts not inScheme).")

    # build parent->children, and broader mapping
    children = defaultdict(list)
    broader = defaultdict(list)

    for c in concepts:
        for p in g.objects(c, SKOS.broader):
            if p in concepts:
                broader[c].append(p)
                children[p].append(c)

    # warn on polyhierarchy
    poly = [c for c, ps in broader.items() if len(ps) > 1]
    if poly:
        print(f"WARNING: {len(poly)} concepts have multiple skos:broader parents; table will pick the first (arbitrary).", file=sys.stderr)

    def sort_nodes(nodes):
        return sorted(nodes, key=lambda c: (code_key(notation(g, c)), label(g, c).lower()))

    tops = sort_nodes(tops)
    for p in list(children.keys()):
        children[p] = sort_nodes(children[p])

    # classify level by counting parts in notation (e.g. 1.0.0 -> 3 parts)
    def parts(code):
        return code.split(".") if code else []

    # Traverse: top -> secondary -> tertiary (3-level expected)
    rows = []  # each row: (p_code, p_label, s_code, s_label, t_code, t_label)
    for top in tops:
        top_code, top_label = notation(g, top), label(g, top)

        secs = children.get(top, [])
        if not secs:
            # allow a top with no children
            rows.append((top_code, top_label, "", "", "", ""))
            continue

        for sec in secs:
            sec_code, sec_label = notation(g, sec), label(g, sec)

            ters = children.get(sec, [])
            if not ters:
                rows.append((top_code, top_label, sec_code, sec_label, "", ""))
                continue

            for ter in ters:
                ter_code, ter_label = notation(g, ter), label(g, ter)
                rows.append((top_code, top_label, sec_code, sec_label, ter_code, ter_label))

    # --- Markdown output (blank repeats to mimic merged cells) ---
    md_lines = []
    md_lines.append("| Primary code | Primary class | Secondary code | Secondary class | Tertiary code | Tertiary class |")
    md_lines.append("|---:|---|---:|---|---:|---|")

    prev_p = prev_s = None
    for p_code, p_lab, s_code, s_lab, t_code, t_lab in rows:
        if (p_code, p_lab) == prev_p:
            p_code_out, p_lab_out = "", ""
        else:
            p_code_out, p_lab_out = p_code, p_lab
            prev_p = (p_code, p_lab)
            prev_s = None  # reset secondary when primary changes

        if s_code == "" and s_lab == "":
            s_code_out, s_lab_out = "", ""
        elif (s_code, s_lab) == prev_s:
            s_code_out, s_lab_out = "", ""
        else:
            s_code_out, s_lab_out = s_code, s_lab
            prev_s = (s_code, s_lab)

        md_lines.append(f"| {p_code_out} | {p_lab_out} | {s_code_out} | {s_lab_out} | {t_code} | {t_lab} |")

    with open(out_md, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines) + "\n")

    # --- XLSX output with real merged cells ---
    wb = Workbook()
    ws = wb.active
    ws.title = "NZLUM"

    headers = ["Primary code", "Primary class", "Secondary code", "Secondary class", "Tertiary code", "Tertiary class"]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EDEDED")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # write rows (no blanking for Excel; we merge instead)
    start_row = 2
    for r in rows:
        ws.append(list(r))

    end_row = start_row + len(rows) - 1

    # helper to merge vertical runs where values repeat
    def merge_runs(col_idx):
        r = start_row
        while r <= end_row:
            v = ws.cell(row=r, column=col_idx).value
            r2 = r
            while r2 + 1 <= end_row and ws.cell(row=r2 + 1, column=col_idx).value == v:
                r2 += 1
            if v not in (None, "") and r2 > r:
                ws.merge_cells(start_row=r, start_column=col_idx, end_row=r2, end_column=col_idx)
                ws.cell(row=r, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)
            r = r2 + 1

    # Merge primary code/class columns (1,2)
    merge_runs(1)
    merge_runs(2)

    # Merge secondary code/class, but only within identical primary groups.
    # Easiest: merge runs on those columns as written (they repeat under same primary anyway)
    merge_runs(3)
    merge_runs(4)

    # formatting
    for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # column widths (tweak)
    widths = [12, 34, 14, 36, 14, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_xlsx)

    print(f"Wrote {out_md} and {out_xlsx}", file=sys.stderr)

if __name__ == "__main__":
    main()
